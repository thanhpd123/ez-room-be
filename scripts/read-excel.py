import openpyxl
import sys

excel_file = r'e:\SEP490\Performance_Test_Report.xlsx'

try:
    print('📖 Reading Excel file...\n')
    
    workbook = openpyxl.load_workbook(excel_file)
    
    print('📊 Workbook Sheets:')
    print(workbook.sheetnames)
    print(f'\nTotal Sheets: {len(workbook.sheetnames)}\n')
    
    # Read each sheet
    for sheet_idx, sheet_name in enumerate(workbook.sheetnames):
        print('\n' + '='*80)
        print(f'Sheet {sheet_idx + 1}: {sheet_name}')
        print('='*80)
        
        worksheet = workbook[sheet_name]
        
        # Get data
        rows_list = []
        for row in worksheet.iter_rows(values_only=True):
            rows_list.append(row)
        
        # Display first 25 rows
        rows_to_show = min(25, len(rows_list))
        print(f'Showing {rows_to_show} of {len(rows_list)} rows:\n')
        
        for i, row in enumerate(rows_list[:rows_to_show]):
            print(f'Row {i + 1}: {row}')
        
        if len(rows_list) > 25:
            print(f'\n... ({len(rows_list) - 25} more rows)')
    
    print('\n' + '='*80)
    print('✅ File read successfully')
    
except Exception as error:
    print(f'❌ Error reading Excel file:')
    print(f'{error}')
    sys.exit(1)
